
# List each company with respective productive count.
# List each company with respective total inventory value. Result: {'AAA Company': 10969059.95, 'BBB Company: 2375499.47', 'CCC Company': 8114363.62}
# List products with inventory less than 10. 
# Write to spreadsheet: Calculate and write inventory value for each product into spreadsheet. 

import openpyxl

file = openpyxl.load_workbook(filename = 'inventory.xlsx')

product_list = file['Sheet1']

products_per_supplier = {}
total_value_per_supplier = {}
product_less_than_value = {}

# range starts with 2: xls product list starts with 2.
# range end with 76: xls product list ends on 75.
# so range must be range(2, 76).
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_no = product_list.cell(product_row, 1).value
    inventory_price_sum_calculation = product_list.cell(product_row, 5)

    # calculation number of products per supplier.
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1
    else:
        products_per_supplier[supplier_name] = 1
    
    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price    
    else:
        total_value_per_supplier[supplier_name] = inventory * price # Last excel row inventory and price. 

    # Products with inventory less than 10
    if inventory < 10:
        product_less_than_value[int(product_no)] = int(inventory)  

    # Calculate and write inventory value
    inventory_price_sum_calculation.value = inventory * price

file.save("inventory_with_total.xlsx")            

print(products_per_supplier) # Result: {'AAA Company': 43, 'BBB Company: 17', 'CCC Company': 14}
print(total_value_per_supplier) # Result: {'AAA Company': 10969059.95, 'BBB Company: 2375499.47', 'CCC Company': 8114363.62}
print(product_less_than_value) # Result: {25: 7, 30: 6, 74: 2}
